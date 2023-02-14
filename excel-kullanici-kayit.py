"""
Bankamatik Arayüz Çalışması
Writer : Murat Can Olgunsoy
Revision Date : 14.02.2023
"""

# Mail gönderebilmek için kullanılan kütüphaneler
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import sys
# Random sayı üretebilmek için kullanılan kütüphane
import random
# Excel dosyası üzerinde işlemler için kullanılan kütüphane
import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.xml.constants import MAX_ROW

# Excel dosyası tanıtılıyor ve işlem yapmaya aktif hale getirilıyor
wb = load_workbook("kullanıcılar.xlsx")
ws = wb.active
ws = wb["Sayfa1"]

# Döngülerde indis belirlemek için bilgilerin depolanacağı diziler oluşturuluyor
kullanıcı_musteriNo = []
kullanıcı_tc = []
kullanıcı_isim = []
kullanıcı_soyisim = []
kullanıcı_doğum = []
kullanıcı_sifre = []
kullanıcı_sifre2 = []
kullanıcı_sifre3 = []
kullanıcı_bakiye = []
kullanıcı_mail = []

# Programın başında ve değişiklerden sonra bilgiler son hali ile dizilere aktarılıyor
def gettingData():

    wb = load_workbook("kullanıcılar.xlsx")
    ws = wb.active
    ws = wb["Sayfa1"]

    # Diziler temizleniyor
    kullanıcı_musteriNo.clear
    kullanıcı_tc.clear 
    kullanıcı_isim.clear 
    kullanıcı_soyisim.clear 
    kullanıcı_doğum.clear 
    kullanıcı_sifre.clear 
    kullanıcı_sifre2.clear 
    kullanıcı_sifre3.clear
    kullanıcı_bakiye.clear 
    kullanıcı_mail.clear
    # Döngü içerisinde bilgilerin başladığı 2. satırdan bulunan maximum satıra kadar bilgiler dizilere aktarılıyor
    for x in range(2,ws.max_row+1):        
        kullanıcı_musteriNo.append(ws.cell(x,1).value)
        kullanıcı_isim.append(ws.cell(x,2).value)
        kullanıcı_soyisim.append(ws.cell(x,3).value)
        kullanıcı_tc.append(ws.cell(x,4).value)
        kullanıcı_sifre.append(ws.cell(x,5).value)
        kullanıcı_sifre2.append(ws.cell(x,6).value)
        kullanıcı_sifre3.append(ws.cell(x,7).value)
        kullanıcı_doğum.append(ws.cell(x,8).value)
        kullanıcı_bakiye.append(ws.cell(x,9).value)
        kullanıcı_mail.append(ws.cell(x,10).value)

""" Program 'tcDogrulama' fonksiyonunun çağırılması ile başlar.
Fonksiyonun içerisinde kullanıcının girdiği tc kimlik numarasının kayıtlı olup olmadığına bakılır."""
def tcDogrulama():

    wb.save("kullanıcılar.xlsx")
    gettingData()
    print("-------------------------------------------------------")
    tc = int(input(" Tc kimlik nurmaranızı giriniz : "))
    # Eğere tc kimlik numarası kayıtlı ise şifre doğrulama aşamasına geçilir
    if tc in kullanıcı_tc:    
        flagSifre = sifreDogrulama(tc)
        # Eğer kullanıcının girdiği şifre ile tc kimlik numarası uyuşuyorsa kullanıcı işlem aşamasına geçilir
        if(flagSifre == 1):
            kullanıcıİşlem()
       
        else:
            cıkıs()
    # Eğer tc kimlik numarası kayıtlı değil ise kayıt yapılıp yapılmayacağı sorulur     
    else:
        print("-------------------------------------------------------")
        flag = int(input(" Daha önceden giriş kaydınız bulunmamaktadır !\nTekrar giriş yapmak istiyorsanız 2'yi, bilgilerinizin kaydedilmesini istiyorsanız 1'i,\nçıkış yapmak istiyorsanız 0'ı tuşlayınız. "))
        if(flag == 2):
            tcDogrulama()
        elif(flag == 1):
            kayıt(tc)        
        elif(flag==0):
            cıkıs()

# Yeni kullanıcı için kayıt yapılacağında bu fonksiyon kullanılır
def kayıt(tc):
    # Kullanıcı bilgileri istenir
        isim = input("İsminizi giriniz : ")
        soyisim = input("Soyisminizi giriniz : ")
        doğum = input("Doğum tarihinizi giriniz (gg/aa/yy) : ")
        mail = input("Mail adresinizi giriniz : ")
        sifre = input("Kendinize bir şifre belirleyiniz : ")
        musteriNo = str(ws.max_row).zfill(10)
        ws.append([musteriNo,isim,soyisim,tc,sifre,True,False,doğum,0,mail])    # Girilen bilgiler dizi olarak çalıştığımız dosyanınn üzerine eklenir    
        wb.save("kullanıcılar.xlsx")             # Bilgilerin girilmesiyle dosya son haliyle kayıt edilir
        selection = int(input("Bilgileriniz kayıt edilmiştir. Tekrar giriş yapmak için 1, çıkış yapmak için 0 tuşlayınız.\n"))    # Kayıt işleminden sonra işlem yapmaya devam edilip edilmeyeceği sorulur
        if (selection == 1):
            tcDogrulama()
        elif(selection == 0):
            cıkıs()
    

# Şifrenin ilgili tc kimlik numarası ile eşleşip eşleşmeyeceğine bu fonksiyonda bakılır
def sifreDogrulama(tc):
    global i
    gettingData()
    sifre = input(" Şifrenizi giriniz : ")     # Kullanıcıdan şifre istenir
    
    uzunluk = len(kullanıcı_tc)        # Tc kimlik numaralarının kayıtlı olduğu dizinin uzunluğu alınır
    i=0
    while i != uzunluk:     # Foksiyona gelen tc kimlik numarasının olduğu satırı bulmak için kayıtlı tc kimlik numarası kadar 'i' indisi döndürülür
        if tc == ws.cell(i+2,4).value:
            
            if sifre == ws.cell(i+2,5).value:
                print("-------------------------------------------------------")
                print("   Hoşgeldin {}, Müşteri Numarası : {}".format(ws.cell(i+2,2).value,ws.cell(i+2,1).value))
                print("-------------------------------------------------------")
                return 1
                break
            else:
                break
        else:
            i = i+1
    print("-------------------------------------------------------")
    #print("\t\Şifre : {}".format(ws.cell(i+2,4).value))

    # Eğer şifre tc kimlik numarasının kayıtlı olduğu satırdaki şifre ile eşleşmediyse aşağıdaki opsiyonlar sorulur
    flagDeneme = int(input(" Şifre ile tc kimlik numaranız uyuşmadı.\nŞifrenizi unuttuysanız 2, tekrar giriş yapmak için 1,\nçıkış yapmak için 0 ı tuşlayınız : "))
    if(flagDeneme==2):
        sifremiUnuttum(tc)
    elif(flagDeneme==1):
        tcDogrulama()
    elif(flagDeneme == 0):        
        return 0

# Şifremi Unuttum fonksiyonunda şifrenin unutulması durumunda şifreyi yenilemek için kayıtlı mail adresine doğrulama kodu gönderilir
def sifremiUnuttum(tc):
    uzunluk = len(kullanıcı_tc)
    b=0
    kod = random.randint(100000,999999) # Random 6 haneli doğrulama kodu oluşturulur
    while b != uzunluk:
        if tc == kullanıcı_tc[b]:
            kullanıcıMail = ws.cell(b+2,10).value
            
            try:               
                        print("-------------------------------------------------------")
                        mail = smtplib.SMTP("smtp.gmail.com",587)
                        mail.ehlo()
                        mail.starttls()
                        mail.login("bankofmco@gmail.com", "359051Mo")

                        mesaj = MIMEMultipart()
                        mesaj["From"] = "bankofmco@gmail.com"           # Gönderen
                        mesaj["To"] = kullanıcıMail                     # Alıcı
                        
                        sifre = ws.cell(b+2,4).value
                        subject = "--- Bank of MCO ---"                                

                        body = " ".join((
                                "Sifre yenileme kodunuz :\n\t\t %s" % kod,
                                ))

                        msg = f'Subject : {subject}\n\n{body}'           
               
                
                        mail.sendmail(mesaj["From"], mesaj["To"], msg.encode("utf-8"))
                        print(" Sistemde bulunan mail adresinize şifreniz başarılı bir şekilde gönderildi.\nGelen kutunuzu veya spam klasörünü kontrol ediniz.")
                        mail.close()
                

                    # Eğer hata olursa hata kodunu yazdırmak için
            except:
                        print("-------------------------------------------------------")
                        print("Hata:", sys.exc_info()[0])
            break
        else:
            b = b+1
    print("-------------------------------------------------------")
    flag = int(input("Mailinize gelen 6 haneli doğruluma kodunu giriniz : "))
    if flag == kod:
        sifreDegistirme()
    else:
        print("-------------------------------------------------------")
        selection = int(input("Gönderilen kod ile eşleştirilemedi !\nKodu tekrar göndermek için 1, çıkış yapmak için 2 tuşlayınız : "))
        if selection == 1:
            sifremiUnuttum()
        elif selection == 2:
            cıkıs()
    
#Kullanıcı ana menüsü oluşturulur.
def kullanıcıİşlem():
    gettingData()
    selection = int(input("\t1 ---> Para çekmek\n\t2 ---> Para yatırmak\n\t3 ---> Şifre Değiştirme\n\t4 ---> Para Transferi\n\t5 ---> Çıkış\nYapmak istediğiniz işlemi giriniz : "))
    if(selection == 1):
        gettingData()
        paraCekme()
        
    elif(selection == 2):
        gettingData()
        paraYatırma()
        
    elif(selection == 3):
        gettingData()
        sifreDegistirme()
    elif(selection == 4):
        gettingData()        
        paraTransfer()

    elif(selection == 5):
        cıkıs()
#Kayıtlı kullanıcılar arası bakiye tranferi yapılır.      
def paraTransfer():    
    transferNo = input("Transfer etmek istediğiniz müşteri numarasını giriniz : ")
    j=0
    while j != len(kullanıcı_tc):
        if transferNo == ws.cell(j+2,1).value:
            miktar = int(input("\tBakiyeniz {}.\nTransfer etmek istediğiniz miktarı giriniz : ".format(ws.cell(i+2,9).value)))
            if miktar <= ws.cell(i+2,9).value:
                ws.cell(j+2,9).value = ws.cell(j+2,9).value + miktar
                ws.cell(i+2,9).value = ws.cell(i+2,9).value - miktar
                wb.save("kullanıcılar.xlsx")
                selection = int(input("{} {} hesabına {} TL transfer edilmiştir.\nGüncel Bakiyeniz : {} TL.\nİşlem yapmaya devam etmek için 1, çıkış yapmak için 0 tuşlayınız : ".format(ws.cell(j+2,2).value,ws.cell(j+2,3).value,miktar,ws.cell(i+2,9).value)))
                if selection == 1:
                    kullanıcıİşlem()
                elif selection == 0:
                    cıkıs()
                
            else:
                selection = int(input("Bakiyeniz yetersiz !\nTekrar denemek için 1. İşlem menüsüne dönmek için 2 tuşlayınız : "))
                if selection == 1:
                    paraTransfer()
                    break
                elif selection == 2:
                    kullanıcıİşlem() 
                    break
                        
        else:
            j = j + 1
    print("Girdiğiniz müşteri numarasına ait kayıt bulunamadı.")
    kullanıcıİşlem()

#Bakiye azaltma uygulaması
def paraCekme():
    print("-------------------------------------------------------")
    print("\tBakiyeniz {} TL.".format(ws.cell((i+2),9).value))
    miktar = int(input("Çekmek istediğiniz mikarı giriniz : "))
    if(miktar <= ws.cell((i+2),9).value):
        ws.cell((i+2),9).value = ws.cell((i+2),9).value - miktar
        wb.save("kullanıcılar.xlsx")
        selection = int(input("Güncel Bakiyeniz : {} TL.\nİşlem yapmaya devam etmek için 1, çıkış yapmak için 0 tuşlayınız : ".format(ws.cell((i+2),9).value)))
        
        if selection == 1:
            kullanıcıİşlem()
        elif selection == 0:
            cıkıs()
    else:
        selection = int(input("Bakiyeniz yetersiz !\nYeniden tutar girmek için 1, çıkış için 0 tuşlayınız : "))
        if(selection == 1):
            paraCekme()
        else:
            cıkıs()
#Bakiyeye ekleme uygulaması
def paraYatırma():
    print("-------------------------------------------------------")
    print("\tBakiyeniz {} TL.".format(ws.cell((i+2),9).value))
    miktar = int(input("Yatırmak istediğiniz mikarı giriniz : "))
    ws.cell((i+2),9).value = ws.cell((i+2),9).value + miktar
    wb.save("kullanıcılar.xlsx") 
    selection = int(input("Hesabınıza {} TL yatırılmıştır.\nGüncel Bakiyeniz : {} TL.\nİşlem yapmaya devam etmek için 1, çıkış yapmak için 0 tuşlayınız : ".format(miktar,ws.cell((i+2),9).value)))
    if selection == 1:
        kullanıcıİşlem()
    elif selection == 0:
        cıkıs()
#Şifre değiştirme ve eski şifrelerle karşılaştırma
def sifreDegistirme():
    print("-------------------------------------------------------")
    #print("Eski şifreleriniz : {}, {} ve {}".format(ws.cell(i+2,4).value,ws.cell(i+2,5).value,ws.cell(i+2,6).value))
    yeniSifre = input("Yeni şifrenizi giriniz : ")
    
    if(ws.cell(i+2,6).value == yeniSifre ):        
        selection = int(input("Yenilemek istediğiniz şifre son 2 şifrenizden farklı olmalıdır.\nYeniden şifre girmek için 1, çıkış yapmak için 0 tuşlayınız : "))
        if(selection == 1):
            sifreDegistirme()
        elif(selection == 0):
            cıkıs()        
    elif (ws.cell(i+2,7).value == yeniSifre):
        selection = int(input("Yenilemek istediğiniz şifre son 2 şifrenizden farklı olmalıdır.\nYeniden şifre girmek için 1, çıkış yapmak için 0 tuşlayınız : "))
        if(selection == 1):
            sifreDegistirme()
        elif(selection == 0):
            cıkıs()
    elif (ws.cell(i+2,5).value == yeniSifre):
        selection = int(input("Yenilemek istediğiniz şifre şu anki şifrenizden farklı olmalıdır.\nYeniden şifre girmek için 1, çıkış yapmak için 0 tuşlayınız : "))
        if(selection == 1):
            sifreDegistirme()
        elif(selection == 0):
            cıkıs()
    else:
        ws.cell(i+2,7).value = ws.cell(i+2,6).value
        ws.cell(i+2,6).value = ws.cell(i+2,5).value
        ws.cell(i+2,5).value = yeniSifre
        wb.save("kullanıcılar.xlsx")
        
        
        selection = int(input("Şifreniz değiştirilmiştir.\nİşlem yapmaya devam etmek için 1, çıkış yapmak için 0 tuşlayınız.\n"))
        if(selection == 1):            
            tcDogrulama()
        elif(selection == 0):
            cıkıs()

def cıkıs():
    print("-------------------------------------------------------")
    wb.save("kullanıcılar.xlsx")
    print("\tÇıkış yapılmıştır! İyi günler dileriz.")
    print("-------------------------------------------------------")
   


"""////////////////////////////////////////////////////////////////////////////////////////"""

print("--------------- Bankamatik Giriş Ekranı ---------------")

tcDogrulama()
